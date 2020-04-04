# Program for copying compound IC50 vals from date spreadsheet into formatted main spreadsheet
# v1-11


# Import libraries
import os
import sys
import re
import csv
import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from compound import Compound


# Get absolute path of a resource
def getAbsPath(relPath):
    try:
        curPath = sys._MEIPASS
    except:
        curPath = os.path.abspath(".")

    return os.path.join(curPath, relPath)


# Global constants
CUR_DIR = os.getcwd()
MONTH_DIR = CUR_DIR + "\\Original IC50\\"
RES_DIR = getAbsPath("res")
COMP_TEMPLATE_SHEET = getAbsPath("Compile_Template.xlsx")
COMP_TEMPLATE_COL_WIDTH = 52
COMP_DATA_COL_WIDTH = 30
IN_DIR = CUR_DIR + "\\Input Files\\"
OUT_DIR = CUR_DIR + "\\Output Files\\"
IC50_ROW = 8
IC50_MAX = 200
IC50_NA = "NA"


# Function main runs the loading of experimental data and its transfer
def main():
    
    # Show tutorial text based on user decision
    # tutorial()

    # Get which mode to run program in
    mode = selectMode()
    
    if mode == "c":

        # Get month(s) to compile data for
        while True:
            # Get .xlsx versions of month files stored in user selected month folder
            month = getMonth()
            monthPath = "{}{}\\".format(MONTH_DIR, month)
            monthFiles = getMonthFiles(monthPath)
            # Compile month files into DFB and DFB_JD files
            numCompiled = compileMonth(month, monthPath, monthFiles)
            print ("Successfully compiled {} files".format(numCompiled))
            # Get user decision on whether to continue compiling months
            if (not moreInFiles()):
                break

    # If "transfer IC50" mode is selected
    elif mode == "t":
        
        # Get output file
        outFileInfo = getFile("o")
        outFileName = outFileInfo[0]
        outFile = outFileInfo[1]
    
        # Get input file(s) and transfer IC50 vals to output file
        while True:
            inFileInfo = getFile("i")
            inFileName = inFileInfo[0]
            inFile = inFileInfo[1]
            # Get IC50 values from all compounds
            compounds = getCompounds(inFile.active)
            inFile.close()
            updateICVals(outFile, outFileName, inFileName, compounds)
            # Get user decision on whether to add more input files
            if (not moreInFiles()):
                break

    input("Program complete. Press <Enter> to exit ")


# Show or do not show tutorial based on user decision
def tutorial():

    # While user input is not valid
    while True:
        showTutorial = input("Show tutorial? (y for yes / n for no): ").strip().lower()
        # If user would like to see tutorial, show it
        if showTutorial == "y":
            print()
            # Open tutorial text file and output each line
            tutorialFile = open("ecp_tutorial.txt")
            for line in tutorialFile:
                print(line)
            print()
            tutorialFile.close()
            break
        elif showTutorial == "n":
            break
        # If user input is something other than y or n, notify user and continue loop
        showError("Input was not y/Y or n/N. Please retry")


# Select which mode to run in
def selectMode():

    mode = "c"
    # While user mode input is not valid
    while True:
        mode = input("Run \"compile month\" or \"transfer IC50\" mode? " +
                     "(c for compile / t for transfer): ").strip().lower()
        # If user selected valid mode
        if mode in ["c", "t"]:
            break
        # If user input was invalid
        showError("Input was not c/C or t/T. Please retry")
    
    return mode


# Get month to compile from user
def getMonth():

    # While user month input is not valid
    while True:
        month = input("Please enter month folder (Month Year): ").strip()
        # If user selected valid month
        if os.path.exists("{}{}\\".format(MONTH_DIR, month)):
            break
        showError("Could not access passed month folder. Please retry")

    return month


# Get files stored in user selected month folder
def getMonthFiles(monthPath):

    monthFiles = []
    # Get month files stored in month folder
    for root, dirs, files in os.walk(monthPath):
        # Iterate through month files and convert any .csv files to .xlsx
        for fileName in files:

            # If current file is not a complete file already
            if "_Homo_" in fileName or "_Het_" in fileName:

                fileStem, fileEnd = fileName.split(".")
                fileStem = "{}{}".format(monthPath, fileStem)
                filePath = "{}{}".format(monthPath, fileName)
                
                # If file ending is .xlsx add to month files list
                if fileEnd == "xlsx":
                    curFile = openpyxl.load_workbook(filePath)
                    monthFiles.append([fileName, curFile])
                
                # If file ending is not .xlsx
                else:
                    # If file ending is .csv, convert file to .xlsx and add to month files list
                    if fileEnd == "csv":
                        # If a .xlsx version does not exist already, create it and add it
                        if not os.path.exists("{}.xlsx".format(fileStem)):
                            curFile = open(filePath)
                            curFile = csvToXlsx(fileStem, curFile)
                            monthFiles.append([fileName, curFile])
                        
                    # If file ending is not .csv, notify user of error
                    else:
                        showFatalError([],
                                       "A file in selected month folder is not .xlsx or .csv")
            
    return monthFiles


# Convert .csv file into .xlsx file
def csvToXlsx(fileStem, csvFile):

    # Create empty .xlsx file
    xlsxFile = openpyxl.Workbook()
    xlsxSheet = xlsxFile.active

    # Copy each line from .csv file to new .xlsx file
    csvRead = csv.reader(csvFile, delimiter = ",")
    for row in csvRead:
        # Remove any illegal characters from file
        row = [ILLEGAL_CHARACTERS_RE.sub("ILLEGAL_CHAR", j) for j in row]
        # Search through row for any strings that can be converted to floats
        for i in range(0, len(row)):
            try:
                row[i] = float(row[i])
            except:
                pass
        
        xlsxSheet.append(row)

    # Save new .xlsx file
    xlsxFilePath = "{}.xlsx".format(fileStem)
    xlsxFile.save(xlsxFilePath)
    
    return openpyxl.load_workbook(xlsxFilePath)


# Compile month IC50 data into two seperate workbook files
def compileMonth(month, monthPath, monthFiles):

    # If there are actually files to be compiled, compile them
    if len(monthFiles) > 0:

        # Create two final files for DFB and DFB_JD
        dfbFile = openpyxl.Workbook()
        dfbSheet = dfbFile.active
        dfb_jdFile = openpyxl.Workbook()
        dfb_jdSheet = dfb_jdFile.active

        # Get compile template file and copy its contents to DFB and DFB_JD files
        templateFile = openpyxl.load_workbook(COMP_TEMPLATE_SHEET)
        templateSheet = templateFile.active
        for r in range(1, templateSheet.max_row + 1):
            for c in range(1, templateSheet.max_column + 1):
                curVal = templateSheet.cell(row = r, column = c).value
                dfbSheet.cell(row = r, column = c).value = curVal
                dfb_jdSheet.cell(row = r, column = c).value = curVal

        # Iterate through month files
        for fileInfo in monthFiles:
            
            curFileName, curFile = fileInfo
            curSheet = curFile.active

            # Select correct output sheet (DFB or DFB_JD) based on current file name
            if "_Homo_" in curFileName:
                outSheet = dfbSheet
            else:
                outSheet = dfb_jdSheet

            # Copy contents of current file to output sheet accounting for offset of newly added data
            columnOffset = outSheet.max_column - 1
            for r in range(1, curSheet.max_row + 1):
                for c in range(2, curSheet.max_column + 1):
                    curVal = curSheet.cell(row = r, column = c).value
                    outCell = outSheet.cell(row = r, column = c + columnOffset)
                    outCell.value = curVal

        # Set widths of columns to be readable
        formatCompSheet(dfbSheet)
        formatCompSheet(dfb_jdSheet)
        
        # Save new DFB and DFB_JD files, if files already open notify user of error
        try:
            # Save files to month folder and close open files
            dfbFile.save("{}{} DFB.xlsx".format(monthPath, month))
            dfb_jdFile.save("{}{} DFB_JD.xlsx".format(monthPath, month))
            closeFiles([templateFile, dfbFile, dfb_jdFile])
        except:
            showFatalError([templateFile, dfbFile, dfb_jdFile],
                           "Problem when saving DFB and DFB_JD files. " +
                           "If files are open, program will fail")

        return len(monthFiles)


# Set widths of columns in file
def formatCompSheet(sheet):

    # Set width of first column (template/key)
    letter = openpyxl.utils.get_column_letter(1)
    sheet.column_dimensions[letter].width = COMP_TEMPLATE_COL_WIDTH

    # Iterate through data columns and set widths
    for c in range(2, sheet.max_column + 1):
        letter = openpyxl.utils.get_column_letter(c)
        sheet.column_dimensions[letter].width = COMP_DATA_COL_WIDTH


# Get input or output sheet
def getFile(fileType):

    curFileName = ""
    curFile = ""
    # Select input or output directory info based on argument
    fileTypeToInfo = {
        "i": ["INPUT", IN_DIR],
        "o": ["OUTPUT", OUT_DIR]
        }
    fileInfo = fileTypeToInfo.get(fileType, "")
    # If argument fileType is not "i" or "o" throw an error and exit program
    if fileInfo == "":
        showFatalError([],
                       "Code is wrong. Argument fileType must be \"i\" or \"o\".")
    
    # Try to open file passed by user until valid name given
    while True:
        try:
            curFileName = input("Please enter {} file name: ".format(fileInfo[0])).strip()
            curFile = openpyxl.load_workbook("{}{}.xlsx".format(fileInfo[1], curFileName))
            break
        except:
            showError("Could not access passed file. Please retry")
            
    return [curFileName, curFile]


# Get experiment data from input file
def getCompounds(inFile):

    compounds = {}
    # Iterate over IC50 vals for each compound
    icVals = inFile[IC50_ROW]
    for colNum in range(2, inFile.max_column + 1):
        
        # Get header of compound and its IC50 val
        colHeader = inFile.cell(row = 1, column = colNum).value.strip().split()
        ic = inFile.cell(row = IC50_ROW, column = colNum).value
        # Get ID of compound and its type
        compoundID = colHeader[0]
        compoundType = colHeader[1]
        
        # If compound's IC50 val contains an "~", set IC50 to approximate val available
        if "~" in str(ic):
            ic = float(re.sub("~|\s", "", ic))
        # If compound's IC50 can be converted to float
        try:
            # If compound's IC50 val greater than 200, set IC50 to "> 200"
            if float(ic) > IC50_MAX:
                ic = "> 200"
        except:
            ic = str(ic).strip()
            pass
        # Add compound's ID, type, and IC50 val to compounds dictionary, with ID as key
        curCompound = Compound(compoundID, compoundType, ic)
        compounds.update([(compoundID, curCompound)])

    return compounds


# Update/set IC50 vals in output file to gotten vals
def updateICVals(outFile, outFileName, date, compounds):

    outFileSheet = outFile.active
    addedCount = 0
    notAdded = list(compounds)
    # Iterate through output file columns until correponding date column reached
    for column in outFileSheet.columns:
        if column[0].value.strip() == date:
            # Iterate through rows in column
            for cell in column:
                rowCompoundID = str(outFileSheet.cell(cell.row, 2).value).strip()
                # Iterate through compound IDs
                for curCompoundID in compounds.keys():
                    # If current compound ID matches row compound ID, set IC50 val
                    if rowCompoundID == curCompoundID:
                        # print ("Added IC50 value of: {}".format(rowCompoundID))
                        addedCount += 1
                        notAdded.remove(rowCompoundID)
                        cell.value = compounds.get(rowCompoundID).getIC()
            break

    # Try to save updated output file, return an error if permission is denied
    try:
        outFile.save("{}{}.xlsx".format(OUT_DIR, outFileName))
    except:
        showFatalError([outFile],
                       "Problem when saving updated output file. " +
                       "If output file is open, program will fail")

    # Print results of val copying
    print ("Successfully added {} IC50 values".format(addedCount))
    print ("Failed to add {} IC50 values".format(len(notAdded)))
    # Print out ID(s) of compounds whose IC50 vals were not copied
    for compoundID in notAdded:
        print ("Failed to add: {}".format(compoundID))


# Get user decision on whether to add more input files
def moreInFiles():
    
    # While user input not valid
    while True:
        print()
        moreInput = input("More input files? (y for yes / n for no): ").strip().lower()
        if moreInput == "y":
            return True
        elif moreInput == "n":
            return False
        # If user input is something other than y or n, notify user and continue loop
        showError("Input was not y/Y or n/N. Please retry")


# Show an error message
def showError(msg):
    
    print ("\n<<< ERROR >>> {}".format(msg))


# Close passed files, show fatal error message and exit program
def showFatalError(filesToClose, msg):

    closeFiles(filesToClose)
    showError(msg)
    input("FATAL: Press <Enter> to exit program ")
    sys.exit()


# Close passed files
def closeFiles(files):

    for file in files:
        file.close()


# Run the main function
main()
